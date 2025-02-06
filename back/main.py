# Librerias
import datetime
import pandas as pd
from fastapi import FastAPI, Form, File, UploadFile, HTTPException
import openpyxl
from pydantic import BaseModel
from typing import List, Optional
import os
from fastapi.responses import HTMLResponse, FileResponse
from fastapi.staticfiles import StaticFiles
from openpyxl import Workbook, load_workbook
from datetime import datetime, date
import zipfile
from starlette.responses import JSONResponse

app = FastAPI()

# Ruta para almacenar los archivos
UPLOAD_DIR = "uploaded_files"
os.makedirs(UPLOAD_DIR, exist_ok=True)

# Ruta del archivo Excel
EXCEL_FILE = os.path.abspath("datos_formulario.xlsx")
EXCEL_DIR = os.path.dirname(EXCEL_FILE)
os.makedirs(EXCEL_DIR, exist_ok=True)

# Crear el archivo Excel si no existe
if not os.path.exists(EXCEL_FILE):
    wb = Workbook()
    ws = wb.active
    ws.title = "Formulario"
    ws.append([
        "NOMBRE COMPLETO", "APELLIDOS 1", "APELLIDO 2", "TIPO DOCUMENTO",
        "NRO DOCUMENTO", "DIRECCION", "FECHA NACIMIENTO", "EDAD",
        "BARRIO", "ESTRATO", "TELEFONO", "CORREO", "PROGRAMA", "SEMESTRE",
        "ARCHIVO DOCUMENTO", "ARCHIVO CIVICA", "ARCHIVO SERVICIOS",
        "ARCHIVO ANEXO 1", "ARCHIVO ANEXO 2"
    ])
    wb.save(EXCEL_FILE)

# Tamaño máximo permitido: 5 MB
MAX_FILE_SIZE = 5 * 1024 * 1024  # 5 MB en bytes

# Validación de los campos del formulario
class FormData(BaseModel):
    NombreCompleto: str
    Apellido1: str
    Apellido2: str
    TipoDocumento: str
    NroDocumento: str
    Direccion: str
    FechaNacimiento: date
    Edad: int
    Barrio: str
    Estrato: int
    Telefono: str
    Correo: str
    Programa: str
    Semestre: int

# Configuración de la ruta para servir los archivos estáticos
app.mount("/static", StaticFiles(directory="../front/static"), name="static")

# Cargar la vista principal
@app.get("/", response_class=HTMLResponse)
async def vista():
    return FileResponse("../front/views/form.html")

# Cargar la vista de administración
@app.get("/administration")
async def vista_admin():
    return FileResponse("../front/views/administration.html")

# Ruta para enviar el formulario al Excel y subir a la carpeta los archivos
@app.post("/formulario/")
async def handle_formulario(
    NombreCompleto: str = Form(...),
    Apellido1: str = Form(...),
    Apellido2: str = Form(...),
    TipoDocumento: str = Form(...),
    NroDocumento: str = Form(...),
    Direccion: str = Form(...),
    FechaNacimiento: date = Form(...),
    Edad: int = Form(...),
    Barrio: str = Form(...),
    Estrato: int = Form(...),
    Telefono: str = Form(...),
    Correo: str = Form(...),
    Programa: str = Form(...),
    Semestre: int = Form(...),
    archivos: List[UploadFile] = File(..., description="Sube exactamente 5 archivos PDF"),
):
    # Validar número exacto de archivos
    if len(archivos) != 5:
        raise HTTPException(status_code=400, detail="Debes enviar exactamente 5 archivos.")

    # Validar que los campos no estén vacíos
    for campo in [
        NombreCompleto, Apellido1, Apellido2, TipoDocumento, NroDocumento, Direccion, Barrio, Telefono, Correo, Programa
    ]:
        if not campo.strip():
            raise HTTPException(status_code=400, detail="Todos los campos son obligatorios.")

    # Validar campos de otros tipos
    if FechaNacimiento is None or Edad <= 0 or Estrato <= 0 or Semestre <= 0:
        raise HTTPException(status_code=400, detail="Todos los campos son obligatorios y deben tener valores válidos.")

    # Guardar los archivos y obtener sus rutas
    nombres_archivos = ["CÉDULA", "CIVICA", "SERVICIOPUBLICOS", "ANEXO1", "ANEXO2"]
    rutas_archivos = []

    # Validar los archivos
    for indice, archivo in enumerate(archivos):
        if not (archivo.filename.endswith(".pdf") or archivo.filename.endswith(".xlsx")):
            raise HTTPException(status_code=400, detail=f"El archivo {archivo.filename} no es un PDF o XLSX válido.")

        # Validar tamaño del archivo
        file_content = await archivo.read()
        if len(file_content) > MAX_FILE_SIZE:
            raise HTTPException(
                status_code=400,
                detail=f"El archivo {archivo.filename} excede el tamaño máximo permitido de 5 MB."
            )

        # Guardar el archivo con un nombre único
        nombre_especifico = nombres_archivos[indice]
        extension = ".xlsx" if nombre_especifico == "ANEXO2" else ".pdf"
        nombre_unico = f"{NroDocumento}_{nombre_especifico}{extension}"
        archivo_path = os.path.join(UPLOAD_DIR, nombre_unico)

        with open(archivo_path, "wb") as f:
            f.write(file_content)

        rutas_archivos.append(archivo_path)

    # Convertir la fecha al formato dd-mm-yyyy
    fecha_convertida = FechaNacimiento.strftime("%d-%m-%Y")

    # Guardar los datos en Excel
    try:
        wb = load_workbook(EXCEL_FILE)
        ws = wb.active

        # Crear la fila con los datos
        fila = [
            NombreCompleto, Apellido1, Apellido2, TipoDocumento, NroDocumento, Direccion, fecha_convertida, Edad, Barrio,
            Estrato, Telefono, Correo, Programa, Semestre
        ] + rutas_archivos
        ws.append(fila)

        wb.save(EXCEL_FILE)
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error al guardar en Excel: {str(e)}")

    #return {"message": "Formulario procesado correctamente."}

    # Leer el contenido del archivo HTML de respuesta exitosa
    try:
        with open("../front/views/endForm.html", "r", encoding="utf-8") as file:
            html_content = file.read()
    except FileNotFoundError:
        raise HTTPException(status_code=404, detail="No se encontró la vista 'endForm.html'.")

    return HTMLResponse(content=html_content, status_code=200)

# Ruta para obtener los datos del Excel
@app.get("/datos/", response_model=List[dict])
async def obtener_datos():
    try:
        wb = load_workbook(EXCEL_FILE)
        ws = wb.active

        # Leer los datos del Excel
        datos = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            datos.append({
                "NombreCompleto": row[0],
                "Apellido1": row[1],
                "Apellido2": row[2],
                "TipoDocumento": row[3],
                "NroDocumento": row[4],
                "Direccion": row[5],
                # Convertir fecha al formato dd-mm-yyyy
                "FechaNacimiento": row[6].strftime("%d-%m-%Y") if isinstance(row[6], datetime) else row[6],
                "Edad": row[7],
                "Barrio": row[8],
                "Estrato": row[9],
                "Telefono": row[10],
                "Correo": row[11],
                "Programa2": row[12],
                "Semestre": row[13],
            })
        # Retornar los datos
        return datos
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error al leer el archivo Excel: {str(e)}")

# Ruta para descargar los archivos asociados a una cédula
@app.get("/descargar/{nro_documento}", response_class=FileResponse)
async def descargar_archivos(nro_documento: str):
    # Filtrar los archivos asociados a la cédula
    archivos_cedula = [
        os.path.join(UPLOAD_DIR, f"{nro_documento}_CÉDULA.pdf"),
        os.path.join(UPLOAD_DIR, f"{nro_documento}_CIVICA.pdf"),
        os.path.join(UPLOAD_DIR, f"{nro_documento}_SERVICIOPUBLICOS.pdf"),
        os.path.join(UPLOAD_DIR, f"{nro_documento}_ANEXO1.pdf"),
        os.path.join(UPLOAD_DIR, f"{nro_documento}_ANEXO2.xlsx"),
    ]

    # Verificar si los archivos existen
    archivos_existentes = [archivo for archivo in archivos_cedula if os.path.exists(archivo)]
    if not archivos_existentes:
        raise HTTPException(status_code=404, detail="No se encontraron documentos para esta cédula.")

    # Crear un archivo zip temporal
    zip_path = os.path.join(UPLOAD_DIR, f"{nro_documento}_documentos.zip")
    with zipfile.ZipFile(zip_path, 'w') as zipf:
        for archivo in archivos_existentes:
            zipf.write(archivo, os.path.basename(archivo))

    # Retornar el archivo zip
    return FileResponse(
        zip_path,
        media_type="application/zip",
        filename=f"{nro_documento}_documentosTiqueteEstudiantil.zip",
    )

# Ruta para descargar el archivo Excel completo
@app.get("/descargar-excel", response_class=FileResponse)
async def descargar_excel():
    if os.path.exists(EXCEL_FILE):
        return FileResponse(
            EXCEL_FILE,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            filename="datos_formulario.xlsx"
        )
    else:
        raise HTTPException(status_code=404, detail="El archivo Excel no existe.")
    
# Ruta para editar los campos del Excel
@app.put("/editar/{nro_documento}")
async def editar_datos(
    nro_documento: str,
    NombreCompleto: str = Form(None),
    Apellido1: str = Form(None),
    Apellido2: str = Form(None),
    TipoDocumento: str = Form(None),
    Direccion: str = Form(None),
    FechaNacimiento: date = Form(None),
    Edad: int = Form(None),
    Barrio: str = Form(None),
    Estrato: int = Form(None),
    Telefono: str = Form(None),
    Correo: str = Form(None),
    Programa2: str = Form(None),
    Semestre: int = Form(None),
):
    try:
        wb = load_workbook(EXCEL_FILE)
        ws = wb.active

        # Buscar la fila correspondiente al número de documento
        for row in ws.iter_rows(min_row=2):
            if str(row[4].value) == nro_documento:  # Asegúrate de comparar como cadenas
                if NombreCompleto is not None:
                    row[0].value = NombreCompleto
                if Apellido1 is not None:
                    row[1].value = Apellido1
                if Apellido2 is not None:
                    row[2].value = Apellido2
                if TipoDocumento is not None:
                    row[3].value = TipoDocumento
                if Direccion is not None:
                    row[5].value = Direccion
                if FechaNacimiento is not None:
                    row[6].value = FechaNacimiento.strftime("%d-%m-%Y")
                if Edad is not None:
                    row[7].value = Edad
                if Barrio is not None:
                    row[8].value = Barrio
                if Estrato is not None:
                    row[9].value = Estrato
                if Telefono is not None:
                    row[10].value = Telefono
                if Correo is not None:
                    row[11].value = Correo
                if Programa2 is not None:
                    row[12].value = Programa2
                if Semestre is not None:
                    row[13].value = Semestre
                break
        else:
            raise HTTPException(status_code=404, detail="No se encontró el documento especificado.")

        wb.save(EXCEL_FILE)
        return {"message": "Datos actualizados correctamente."}
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error al actualizar los datos: {str(e)}")
    
@app.get("/download-file/{filename}")
def download_file(filename: str):
    """
    Permite descargar un archivo específico.
    """
    file_path = os.path.join(UPLOAD_DIR, filename)
    if os.path.exists(file_path):
        return FileResponse(file_path, media_type='application/pdf', filename=filename)
    return {"error": "File not found"}

# Endpoint para cargar un archivo (reemplaza el archivo anterior si existe)
@app.post("/upload-file/{nro_documento}")
def upload_file(nro_documento: str, file: UploadFile = File(...)):
    """
    Sube un archivo PDF para un usuario basado en su número de documento.
    """
    file_extension = os.path.splitext(file.filename)[1]
    if file_extension.lower() not in [".pdf",".xlsx"]:
        raise HTTPException(status_code=400, detail="Solo se permiten archivos PDF o XLSX en el Anexo 2.")
    
    file_path = os.path.join(UPLOAD_DIR, file.filename)  # Mantiene el mismo nombre
    
    with open(file_path, "wb") as buffer:
        buffer.write(file.file.read())
    
    return JSONResponse(content={"message": f"Archivo {file.filename} subido exitosamente."})


# Lista de tipos de archivo
nombres_archivos = ["CEDULA", "CIVICA", "SERVICIOSPUBLICOS", "ANEXO1", "ANEXO2"]

# Función para generar el nombre del archivo basado en el número de documento y tipo
# def generar_nombre_archivo(nro_documento: str, tipo: str) -> str:
#     return f"{nro_documento}_{tipo}.pdf"

# Función auxiliar para eliminar el archivo si existe
def eliminar_archivo_si_existe(file_path):
    if os.path.exists(file_path):
        os.remove(file_path)
        return True
    return False


# Función que obtiene los archivos asociados a un número de documento
def obtener_archivos_por_documento(nro_documento: str):
    archivos = []
    # Supongamos que los archivos tienen un formato como {nro_documento}_CÉDULA.pdf, etc.
    for archivo in os.listdir(UPLOAD_DIR):
        if archivo.startswith(nro_documento) and (archivo.endswith(".pdf") or archivo.endswith(".xlsx")):
            archivos.append(archivo)
    return archivos

@app.get("/mostrar-archivos/{nro_documento}")
async def mostrar_archivos(nro_documento: str):
    archivos = obtener_archivos_por_documento(nro_documento)
    if archivos:
        return JSONResponse(content={"files": archivos})
    else:
        return JSONResponse(content={"files": []})  # Retornar una respuesta vacía si no hay archivos
    
# Ruta para listar los archivos PDF de un usuario específico
@app.get("/list-files/{cedula}")
def list_files(cedula: str):
    """
    Lista los archivos PDF de un usuario específico según la cédula.
    """
    files = []
    if os.path.exists(UPLOAD_DIR):
        files = [
            file for file in os.listdir(UPLOAD_DIR) 
            if file.startswith(cedula) and (file.endswith(".pdf") or file.endswith(".xlsx"))
        ]
    return {"files": files}

# Endpoint para eliminar un archivo específico
@app.delete("/delete-file/{cedula}/{filename}")
def delete_file(cedula: str, filename: str):
    """
    Elimina un archivo PDF específico de un usuario según la cédula y el nombre del archivo.
    """
    # Generar la ruta completa del archivo
    file_path = os.path.join(UPLOAD_DIR, filename)
    
    # Eliminar el archivo si existe
    if eliminar_archivo_si_existe(file_path):
        return {"message": f"Archivo {filename} eliminado exitosamente."}
    else:
        raise HTTPException(status_code=404, detail=f"Archivo {filename} no encontrado.")

# Eliminar registro y archivos pdf asociados
@app.delete("/eliminar/{document_id}")
async def eliminar_registro(document_id: str):
    # Cargar el archivo Excel
    if not os.path.exists(EXCEL_FILE):
        raise HTTPException(status_code=404, detail="Archivo Excel no encontrado")

    df = pd.read_excel(EXCEL_FILE)

    # Verificar si el documento existe en el archivo Excel
    if 'NRO DOCUMENTO' not in df.columns:
        raise HTTPException(status_code=404, detail="Columna 'NRO DOCUMENTO' no encontrada en el archivo Excel")

    if document_id not in df['NRO DOCUMENTO'].astype(str).values:
        raise HTTPException(status_code=404, detail="Documento no encontrado en el archivo Excel")

    # Eliminar el registro del DataFrame
    df = df[df['NRO DOCUMENTO'].astype(str) != document_id]

    # Guardar el DataFrame actualizado en el archivo Excel
    df.to_excel(EXCEL_FILE, index=False)

    # Eliminar los archivos PDF correspondientes en la carpeta uploaded_files
    # pdf_files = [f for f in os.listdir(UPLOAD_DIR) if f.startswith(document_id)]
    # for pdf_file in pdf_files:
    #     os.remove(os.path.join(UPLOAD_DIR, pdf_file))

    # return {"detail": "Registro eliminado correctamente"}
    # Eliminar los archivos PDF y Excel correspondientes en la carpeta uploaded_files
    files_to_delete = [f for f in os.listdir(UPLOAD_DIR) if f.startswith(document_id) and (f.endswith(".pdf") or f.endswith(".xlsx"))]
    for file in files_to_delete:
        os.remove(os.path.join(UPLOAD_DIR, file))

    return {"detail": "Registro eliminado correctamente"}
